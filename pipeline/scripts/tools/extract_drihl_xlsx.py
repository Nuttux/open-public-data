#!/usr/bin/env python3
"""
Extracteur DRIHL — Socle de données demandes et attributions de logements sociaux.

Input  : XLSX DRIHL "Socle de données demandes et attributions {YYYY}"
         https://www.drihl.ile-de-france.developpement-durable.gouv.fr/
         socle-de-donnees-demandes-et-attributions-de-a1414.html

Output : pipeline/seeds/seed_drihl_paris_{YYYY}.csv
         - 20 lignes arrondissements Paris (INSEE 75101-75120)
         - 1 ligne Paris global (INSEE 75056)
         - Colonnes: code_insee, nom, demandes_choix1, attributions,
                    delai_median_mois, part_anciennete_5ans_plus, ratio_dem_attrib

Usage:
    python scripts/tools/extract_drihl_xlsx.py \\
        --input /path/to/socle_demandes_attributions_2024.xlsx \\
        --year 2024

    # ou avec téléchargement automatique:
    python scripts/tools/extract_drihl_xlsx.py --year 2024 --download
"""

from __future__ import annotations

import argparse
import csv
import re
import urllib.request
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent.parent.parent
SEED_DIR = REPO_ROOT / "pipeline" / "seeds"

# Columns found in DRIHL "Ensemble" sheet (row 3 is the header)
# Indices are 0-based after openpyxl returns tuple
COL_NIV_GEO = 0       # "Région" / "Département" / "Commune" / ...
COL_CODE = 1          # INSEE code (commune: 75056 for Paris, 75101-75120 for arr)
COL_NOM = 2
COL_DEMANDES_TOT = 9  # "Nombre total de demandes ciblant le territoire"
COL_DEMANDES_CHOIX1 = 10  # "Nombre de demandes ciblant le territoire en choix 1"
COL_ATTRIBUTIONS = 13  # "Nombre d'attributions"
COL_DELAI_MEDIAN_MOIS = 18  # "Délai médian d'attribution (en mois)"

# Ancienneté breakdown columns (to compute "5 ans et plus")
# We'll locate them dynamically by header match

DEFAULT_URL_TPL = (
    "https://www.drihl.ile-de-france.developpement-durable.gouv.fr/"
    "IMG/xlsx/socle_demandes_attributions_{year}.xlsx"
)


def _locate_anciennete_cols(headers: list) -> dict[str, int]:
    """Find column indices for ancienneté breakdown in 'Ensemble' sheet."""
    out: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        hs = str(h).strip().lower().replace("\n", " ")
        if re.search(r"5\s*ans.*plus|au moins 5 ans|depuis\s+5", hs):
            out.setdefault("part_anciennete_5ans_plus", i)
        elif re.search(r"moins.*1\s*an", hs):
            out.setdefault("part_anciennete_moins_1an", i)
    return out


def extract(xlsx_path: Path, year: int) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Ensemble"]

    # Header row is row 3 (1-indexed)
    headers = list(next(ws.iter_rows(min_row=3, max_row=3, values_only=True)))
    anc_cols = _locate_anciennete_cols(headers)

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        niv = row[COL_NIV_GEO]
        code = row[COL_CODE]
        nom = row[COL_NOM]
        if code is None:
            continue
        code_str = str(code).strip()

        # Keep:
        #  - Département Paris (75)
        #  - Communes Paris arrondissements (75101..75120) + Paris non précisé (75056)
        is_paris_dept = (
            str(niv).strip().lower() == "département" and code_str == "75"
        )
        is_paris_commune = (
            str(niv).strip().lower() == "commune"
            and code_str.startswith("751")
            and len(code_str) == 5
        )
        is_paris_commune_generic = (
            str(niv).strip().lower() == "commune" and code_str == "75056"
        )
        if not (is_paris_dept or is_paris_commune or is_paris_commune_generic):
            continue

        demandes_choix1 = row[COL_DEMANDES_CHOIX1]
        attributions = row[COL_ATTRIBUTIONS]
        delai_median = row[COL_DELAI_MEDIAN_MOIS]
        part_5ans = (
            row[anc_cols["part_anciennete_5ans_plus"]]
            if "part_anciennete_5ans_plus" in anc_cols
            else None
        )

        def _num(v):
            if v is None or v == "-":
                return None
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        dem = _num(demandes_choix1)
        att = _num(attributions)
        ratio = (dem / att) if (dem and att and att > 0) else None

        rows.append(
            {
                "code_insee": code_str,
                "nom": (nom or "").strip(),
                "niveau_geo": ("departement" if is_paris_dept else "commune"),
                "annee": year,
                "demandes_choix1": int(dem) if dem else None,
                "attributions": int(att) if att else None,
                "ratio_dem_attrib": round(ratio, 2) if ratio else None,
                "delai_median_attribution_mois": (
                    round(float(delai_median), 1)
                    if isinstance(delai_median, (int, float))
                    else None
                ),
                "part_anciennete_5ans_plus": (
                    round(float(part_5ans), 4)
                    if isinstance(part_5ans, (int, float))
                    else None
                ),
                "source": "DRIHL - Socle de données demandes et attributions",
                "source_url": (
                    "https://www.drihl.ile-de-france.developpement-durable.gouv.fr/"
                    "socle-de-donnees-demandes-et-attributions-de-a1414.html"
                ),
            }
        )
    return rows


def write_seed(rows: list[dict], year: int) -> Path:
    out = SEED_DIR / f"seed_drihl_paris_{year}.csv"
    cols = [
        "code_insee",
        "nom",
        "niveau_geo",
        "annee",
        "demandes_choix1",
        "attributions",
        "ratio_dem_attrib",
        "delai_median_attribution_mois",
        "part_anciennete_5ans_plus",
        "source",
        "source_url",
    ]
    with out.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, quoting=csv.QUOTE_MINIMAL)
        w.writeheader()
        for r in rows:
            w.writerow({c: (r.get(c) if r.get(c) is not None else "") for c in cols})
    return out


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", type=Path, help="Chemin vers le XLSX DRIHL local")
    p.add_argument("--year", type=int, required=True, help="Année du socle (ex: 2024)")
    p.add_argument(
        "--download",
        action="store_true",
        help=f"Télécharge depuis {DEFAULT_URL_TPL}",
    )
    args = p.parse_args()

    if args.download:
        url = DEFAULT_URL_TPL.format(year=args.year)
        tmp = Path(f"/tmp/drihl_{args.year}.xlsx")
        print(f"↓ Téléchargement {url}")
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req) as resp, tmp.open("wb") as f:
            f.write(resp.read())
        xlsx_path = tmp
    elif args.input:
        xlsx_path = args.input
    else:
        raise SystemExit("Spécifier --input PATH ou --download")

    rows = extract(xlsx_path, args.year)
    out = write_seed(rows, args.year)
    print(f"✓ {len(rows)} lignes → {out}")
    # Quick preview
    for r in rows[:5] + rows[-3:]:
        print(
            f"  {r['code_insee']:>5s} | {r['nom'][:35]:35s} | "
            f"dem={r['demandes_choix1'] or '-':>7} | "
            f"attr={r['attributions'] or '-':>5} | "
            f"ratio={r['ratio_dem_attrib'] or '-':>6} | "
            f"delai={r['delai_median_attribution_mois'] or '-':>5}m"
        )


if __name__ == "__main__":
    main()
